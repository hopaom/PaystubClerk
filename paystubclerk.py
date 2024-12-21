import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import sys
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Protection
from openpyxl.utils import range_boundaries, get_column_letter
from copy import copy
from datetime import datetime, timedelta
from PIL import Image, ImageTk

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class PaystubClerk:
    def __init__(self, master):
        self.master = master
        master.title("PaystubClerk_Ho")
        master.geometry("550x850")
        self.input_file_path = tk.StringVar()
        self.output_directory = tk.StringVar()
        self.create_widgets()
        self.load_image()

    def load_image(self):
        image_path = resource_path("manual.png")  # 설명할 이미지 경로를 입력
        image = Image.open(image_path)
        image = image.resize((396, 549))  # 이미지 크기 조정
        self.photo = ImageTk.PhotoImage(image)
        self.image_label = tk.Label(self.master, image=self.photo)
        self.image_label.grid(row=6, column=0, columnspan=3, pady=10)

    def create_widgets(self):
        tk.Label(self.master, text="Input File:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        tk.Entry(self.master, textvariable=self.input_file_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(self.master, text="Browse", command=self.browse_input_file).grid(row=0, column=2, padx=5, pady=5)

        tk.Label(self.master, text="Output Directory:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        tk.Entry(self.master, textvariable=self.output_directory, width=50).grid(row=1, column=1, padx=5, pady=5)
        tk.Button(self.master, text="Browse", command=self.browse_output_directory).grid(row=1, column=2, padx=5, pady=5)

        tk.Label(self.master, text="Column Positions:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.create_column_inputs()
        
        tk.Button(self.master, text="Convert", command=self.convert).grid(row=4, column=0, columnspan=3, pady=10)
        
        self.status_message = tk.StringVar()
        tk.Label(self.master, textvariable=self.status_message).grid(row=5, column=0, columnspan=3, pady=5)

    def create_column_inputs(self):
        column_frame = ttk.Frame(self.master)
        column_frame.grid(row=3, column=0, columnspan=3, padx=5, pady=5, sticky="w")
        
        columns = [
            ("Name", 2),
            ("Position", 1),
            ("Work Type", 4),
            ("Work Hours", 5),
            ("Hourly Rate", 6),
            ("Gross Pay", 7),
            ("Income Tax", 8)
        ]
        
        for i, (label, default) in enumerate(columns):
            tk.Label(column_frame, text=f"{label}:").grid(row=i//3, column=(i%3)*2, sticky="e", padx=5, pady=2)
            entry = tk.Entry(column_frame, width=5)
            entry.insert(0, str(default))
            entry.grid(row=i//3, column=(i%3)*2+1, padx=5, pady=2)
            setattr(self, f"{label.lower().replace(' ', '_')}_column", entry)

    def browse_input_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.input_file_path.set(filename)

    def browse_output_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_directory.set(directory)

    def convert(self):
        if not self.input_file_path.get():
            messagebox.showerror("Error", "Please select an input file.")
            return
        if not self.output_directory.get():
            messagebox.showerror("Error", "Please select an output directory.")
            return
        
        try:
            self.process_payslips()
            self.status_message.set("Conversion completed successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during conversion: {str(e)}")

    def process_payslips(self):
        input_file_path = self.input_file_path.get()
        output_directory = self.output_directory.get()
        
        wb = load_workbook(input_file_path, data_only=True)
        sheet = wb.active
        merged_cell_map = self.create_merged_cell_map(sheet)
        
        template_sheet = wb["Template"]
        
        pay_date_cell = sheet.cell(row=1, column=1).value
        pay_date = None
        
        if isinstance(pay_date_cell, str):
            pay_date = datetime.strptime(pay_date_cell,"%Y-%m-%d").replace(day=1) + timedelta(days=32)
            pay_date = pay_date.replace(day=10)
        
        elif isinstance(pay_date_cell,datetime):
            pay_date = pay_date_cell.replace(day=1) + timedelta(days=32)
            pay_date = pay_date.replace(day=10)

        name_col = int(self.name_column.get())
        position_col = int(self.position_column.get())
        work_type_col = int(self.work_type_column.get())
        work_hours_col = int(self.work_hours_column.get())
        hourly_rate_col = int(self.hourly_rate_column.get())
        gross_pay_col = int(self.gross_pay_column.get())
        income_tax_col = int(self.income_tax_column.get())

        current_person = None
        work_types_list = []
        work_hours_list = []
        hourly_rates_list = []
        gross_pays_list = []
        income_tax_list = []
        
        current_position = ""

        for row_index in range(3,sheet.max_row + 1):
            name = self.get_cell_value(sheet,row_index,name_col ,merged_cell_map)
            
            if not name:
                continue
            
            position = self.get_cell_value(sheet,row_index ,position_col ,merged_cell_map)
            work_type = self.get_cell_value(sheet,row_index ,work_type_col ,merged_cell_map)
            work_hours = self.get_cell_value(sheet,row_index ,work_hours_col ,merged_cell_map)
            hourly_rate = self.get_cell_value(sheet,row_index ,hourly_rate_col ,merged_cell_map)
            gross_pay = self.get_cell_value(sheet,row_index ,gross_pay_col ,merged_cell_map)
            income_tax = self.get_cell_value(sheet,row_index ,income_tax_col ,merged_cell_map)

            if current_person != name and current_person is not None:
                wb_out = self.create_payslip(current_person,current_position,
                                              work_types_list,
                                              work_hours_list,
                                              hourly_rates_list,
                                              gross_pays_list,
                                              income_tax_list,
                                              pay_date,
                                              template_sheet)

                excel_path=os.path.join(output_directory,f"{current_person}_급여명세서.xlsx")
                wb_out.save(excel_path)

                work_types_list=[]
                work_hours_list=[]
                hourly_rates_list=[]
                gross_pays_list=[]
                income_tax_list=[]

            current_person=name
            current_position=position

            if work_hours and hourly_rate:
                work_types_list.append(work_type)
                work_hours_list.append(work_hours)
                hourly_rates_list.append(hourly_rate)
                gross_pays_list.append(gross_pay)
                income_tax_list.append(income_tax)

        
        if current_person:
            wb_out=self.create_payslip(current_person,current_position,
                                        work_types_list,
                                        work_hours_list,
                                        hourly_rates_list,
                                        gross_pays_list,
                                        income_tax_list,
                                        pay_date,
                                        template_sheet)

            excel_path=os.path.join(output_directory,f"{current_person}_급여명세서.xlsx")
            wb_out.save(excel_path)

    def create_merged_cell_map(self,sheet):
        merged_cell_map={}
        
        for merged_range in sheet.merged_cells.ranges:
            min_col,min_row,max_col,max_row = range_boundaries(str(merged_range))
            
            value = sheet.cell(min_row,min_col).value
            
            for row in range(min_row,max_row + 1):
                for col in range(min_col,max_col + 1):
                    merged_cell_map[(row,col)] = value
                    
                    
                    
                    
                    return merged_cell_map

    def get_cell_value(self,sheet,row,col ,merged_cell_map):
         return merged_cell_map.get((row,col),sheet.cell(row,col).value)

    def create_payslip(self,name ,position ,
                       work_types_list ,
                       work_hours_list ,
                       hourly_rates_list ,
                       gross_pays_list ,
                       income_tax_list ,
                       pay_date ,
                       template_sheet):

         wb_out = Workbook()
         ws_out = wb_out.active
         ws_out.title ="급여명세서"

         for row in template_sheet.iter_rows():
             for cell in row:
                 new_cell = ws_out.cell(row=cell.row,column=cell.column,value=cell.value)

                 if cell.has_style:
                     new_cell.font  = copy(cell.font )
                     new_cell.border  = copy(cell.border )
                     new_cell.fill  = copy(cell.fill )
                     new_cell.number_format  = cell.number_format
                     new_cell.alignment  = copy(cell.alignment )
                     new_cell.protection  = Protection(locked=
                                cell.protection.locked,
                                hidden=
                                cell.protection.hidden )

         for merged_range in template_sheet.merged_cells.ranges:
             min_col,min_row,max_col,max_row = range_boundaries(str(merged_range))
             ws_out.merge_cells(start_row=min_row,start_column=min_col,end_row=max_row,end_column=max_col)

         ws_out["B4"]=name
         ws_out["D4"]=pay_date.strftime("%Y-%m-%d") if pay_date else ""
         ws_out["B5"]=position

         total_gross_pay=sum(gross_pays_list)
         total_income_tax=sum(income_tax_list)
         net_pay=(total_gross_pay - total_income_tax)

         ws_out["B8"]=total_gross_pay
         ws_out["B11"]=total_gross_pay
         ws_out["D8"]=total_income_tax
         ws_out["D10"]=total_income_tax
         ws_out["D11"]=net_pay

         start_row = 15

         for i,(wt ,wh ,hr ,gp) in enumerate(zip(work_types_list ,
                                                  work_hours_list ,
                                                  hourly_rates_list ,
                                                  gross_pays_list)):
             ws_out[f"A{start_row+i}"]=wt if wt else "기본급"
             ws_out[f"B{start_row+i}"]=wh
             ws_out[f"C{start_row+i}"]=hr
             ws_out[f"D{start_row+i}"]=gp

         for col in range(1 ,ws_out.max_column + 1):
             column_letter=get_column_letter(col )
             ws_out.column_dimensions[column_letter].width = 19

         for row in range(1 ,ws_out.max_row + 1):
             ws_out.row_dimensions[row].height = 21

         return wb_out

if __name__ == "__main__":
    root=tk.Tk()
    app=PaystubClerk(root)
    root.mainloop()
